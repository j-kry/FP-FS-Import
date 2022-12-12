import tkinter as tk
from tkinter.filedialog import askopenfilename
import openpyxl

from openpyxl import Workbook
from openpyxl import load_workbook

import requests

import json

import time

#Will store data from csv in a Ticket Object
class Ticket:

    def __init__(self, subject, description, newCat, newSubCat, pendingReason):
        self.subject = subject
        self.description = description
        self.newCat = newCat
        self.newSubCat = newSubCat
        self.pendingReason = pendingReason
    
    def __repr__(self):
        return self.subject + "\n" + self.description

    def getSubject(self):
        return self.subject
    def getDescription(self):
        return self.description
    def getNewCat(self):
        return self.newCat
    def getNewSubCat(self):
        return self.newSubCat
    def getPendingReason(self):
        return self.pendingReason

    def setSubject(self, subject):
        self.subject = subject
    def setDescription(self, description):
        self.description = description
    def setNewCat(self, newCat):
        self.newCat = newCat
    def setNewSubCat(self, newSubCat):
        self.newSubCat = newSubCat
    def setPendingReason(self, pendingReason):
        self.pendingReason = pendingReason

def FileOpen():

    #Path = the one from the window
    filePath = askopenfilename(filetypes=[("Microsoft Excel Worksheet", ".xlsx"), ("All Files", "*.*")])

    if not filePath:
        print("No file present or none selected!")
        return

    return filePath 

##################
#####CONSTANTS####
##################

#Phil ID
REQUESTER_ID = 20001321742
#Justin-Test ID
JUSTIN_REQUESTER_ID = 20000651361
#Phil Group EMR Billing
BILLING_GROUP_ID = 20000342334
#HIS Group
HIS_GROUP_ID = 20000341837
#Justin-Test Group ID
JUSTIN_GROUP_ID = 20000342352

#pending
STATUS = 3
#low
PRIORITY = 1
#slack, not to muddy reports
SOURCE = 10
#hidden category
CATEGORY = "Imported Billing Ticket"
#tags need to be an array
TAGS = ["Imported Billing Ticket"]

JSONHEADER = {'Content-Type' : 'application/json'}

####################
#####Begin MAIN#####
####################

# Load workbook from file prompt
wb = load_workbook(filename=FileOpen())

keyfile = open(FileOpen(), 'r')
APIKEY = keyfile.readline()

# Load worksheet
sheet = wb.active

# Get number of tickets in spreadsheet
numTickets = sheet.max_row

print("READ " + str(numTickets-1) + " TICKETS FROM EXCEL")

#tickets array will hold all Ticket Objects
tickets = []

#Iterate through the sheet starting at the second row
#numTickets-1 because we are adding 2 to the row. Loop range is not inclusive.
for row in range(numTickets-1):

    subject = "IMPORTED-" + sheet.cell(row=row+2, column=9).value
    tikNum = sheet.cell(row=row+2, column=2).value
    createdOn = str(sheet.cell(row=row+2, column=5).value)
    reqReason = sheet.cell(row=row+2, column=10).value
    reqArea = sheet.cell(row=row+2, column=11).value
    reqSubType = sheet.cell(row=row+2, column=12).value
    origSubmitter = sheet.cell(row=row+2, column=17).value + " " + sheet.cell(row=row+2, column=18).value
    jobTitle = sheet.cell(row=row+2, column=21).value
    team = sheet.cell(row=row+2, column=22).value
    cc = sheet.cell(row=row+2, column=29).value
    attachName = sheet.cell(row=row+2, column=32).value
    desc = sheet.cell(row=row+2, column=15).value
    newCat = sheet.cell(row=row+2, column=13).value
    newSubCat = sheet.cell(row=row+2, column=14).value
    pendReason = sheet.cell(row=row+2, column=34).value

    tickets.append(
        Ticket(
        subject, 
        f"Original ticket number: {tikNum}<br>" +
        f"Created on: {createdOn}<br>" +
        f"Reason for request: {reqReason}<br>" +
        f"Request area: {reqArea}<br>" +
        f"Request sub-type: {reqSubType}<br>" +
        f"Originally submitted by: {origSubmitter}<br>" +
        f"Job Title: {jobTitle}<br>" +
        f"Team: {team}<br>" +
        f"CC: {cc}<br>" +
        f"Attachment names: {attachName}<br>" +
        f"Original Description: {desc}<br>",
        newCat,
        newSubCat,
        pendReason)
        )

# for i in tickets:
#     print(i)

counter = 1

for j in tickets:

    subcategory = "" if j.getNewSubCat() == "None" else j.getNewSubCat()

    #Test payload
    #Source is Slack to differentiate from 'real' tickets in reports
    payloadNew = {'requester_id': JUSTIN_REQUESTER_ID, 
    'group_id' : HIS_GROUP_ID, 
    'subject' : j.getSubject(),
    'status' : STATUS,
    'priority' : PRIORITY,
    'description' : j.getDescription(),
    'source' : SOURCE,
    'category' : j.getNewCat(),
    'sub_category': subcategory,
    'custom_fields' : {"pending_reason" : j.getPendingReason()},
    'tags' : TAGS
    }

    try:
        print(requests.post(
            'https://thresholds.freshservice.com/api/v2/tickets',
            json=payloadNew, 
            headers=JSONHEADER,
            auth=(APIKEY, "X")
            ).raise_for_status())
    except requests.exceptions.HTTPError as err:
        print(err)

    print("SENDING TICKET " + str(counter) + " of " + str(numTickets-1))
    counter+=1

    #Wait a second every other ticket to avoid API rate limit
    if(counter%2 == 0):
        time.sleep(1)